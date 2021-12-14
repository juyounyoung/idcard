from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.core.paginator import Paginator
from django.core.files.storage import FileSystemStorage
from urllib.parse import quote
from django.db.models import Q
import urllib
import mimetypes
from mimetypes import guess_type
from django.contrib import messages
from django.urls import reverse
from django.views.generic import ListView
import pandas as pd
import datetime
import random
import os

from .models import *
from openpyxl import load_workbook

from django.contrib.auth.forms import PasswordChangeForm
from .forms import CustomPasswordChangeForm
from django.contrib.auth import update_session_auth_hash


# Create your views here.
# @login_message_required
def login(request):
    if request.method == 'POST':
        username2 = request.POST['school_id']
        password2 = request.POST['password']

        if Users_user.objects.filter(school_ID=username2).exists():
            getUser = Users_user.objects.get(school_ID=username2)
            # 로그인 성공 -> board_list
            if getUser.password == password2 and getUser.login_trial != 5:
                getName = school_info.objects.get(school_ID=username2)
                request.session['school_name'] = getName.school_name
                request.session['school_ID'] = username2
                # 로그인시도횟수 초기화
                getUser.login_trial = 0
                getUser.save()
                # 첫번째 로그인 -> 비밀번호 변경 팝업
                if password2 == 'password1':
                    # return render(request, 'pw_edit.html')
                    return render(request, 'login.html', {'pw_edit': 1})
                else:
                    return redirect('/board')
            # 로그인 실패 -> 로그인시도횟수 추가
            else:
                chklogin = ''
                if getUser.login_trial == 5:
                    chklogin = '로그인 시도 횟수 5회 초과하여 로그인할 수 없습니다. 관리자에게 문의하세요'
                else:
                    getUser.login_trial += 1
                    getUser.save()
                    chklogin = '로그인 시도 횟수 : ' + str(getUser.login_trial) + '회'
                messages.warning(request, chklogin)
                return render(request, 'login.html')
    return render(request, 'login.html')


# 비밀번호 변경 팝업창(미완성)
def pw_edit(request):
    if request.method == 'POST':
        password_change_form = CustomPasswordChangeForm(request, request.POST)
        context = {'password_change_form': password_change_form}
        # if password_change_form.is_valid():
        school_id = request.session['school_ID']
        old_password = request.POST['old_password']
        new_password1 = request.POST['new_password1']
        new_password2 = request.POST['new_password2']

        getUser = Users_user.objects.get(school_ID=school_id)
        # 기존 비밀번호
        if getUser.password == old_password:
            if len(new_password1) < 8:
                messages.warning(request, '최소 길이 에러')
                return render(request, 'pw_edit.html', context)
            if new_password1 == new_password2:
                getUser.password = new_password1
                getUser.save()
            else:
                messages.warning(request, '새 비밀번호를 다시 확인해주세요')
                return render(request, 'pw_edit.html', context)
        else:
            messages.warning(request, '기존 비밀번호를 다시 확인해주세요')
            return render(request, 'pw_edit.html', context)

        messages.success(request, "비밀번호를 성공적으로 변경하였습니다.")
        return render(request, 'pw_edit.html', context)
    else:
        password_change_form = CustomPasswordChangeForm(request)

    return render(request, 'pw_edit.html', {'password_change_form': password_change_form})


# 게시판 조회
def board(request):
    # 페이징 10개씩
    all_boards = Board.objects.all()  # 정렬 기준 추가 필요
    paginator = Paginator(all_boards, 10)
    page = int(request.GET.get('page', 1))
    boards = paginator.get_page(page)

    search_keyword = request.GET.get('search', '')
    if search_keyword:
        search_board_list = all_boards.filter(title__icontains=search_keyword)
        paginator = Paginator(search_board_list, 10)
        page = int(request.GET.get('page', 1))
        search_board_list = paginator.get_page(page)
        return render(request, 'board_list.html', {'boards': search_board_list, 'search': search_keyword})

    return render(request, 'board_list.html', {'boards': boards})


def board_write(request):
    if Board.objects.all().exists():
        obj = Board.objects.latest('id')
        new_board_id = int(obj.id) + 1
    else:
        new_board_id = 1
    post_id = request.GET.get('board_id', new_board_id)
    school_id = request.session['school_ID']
    school_name = request.session['school_name']

    # 학교 당 게시글 하나
    if school_id != 'admin' and Board.objects.filter(school_name=school_name).exists():
        messages.warning(request, '게시글은 하나씩만 작성 가능합니다')
        return redirect('/board')

    slist = []

    if request.FILES.get('excelfile') is not None:
        print("excel not none")
        excel_file = request.FILES.get('excelfile')
        fs = FileSystemStorage(location='static/board/xlsx')
        name = fs.save(excel_file.name, excel_file)

        sheets = load_workbook(excel_file, read_only=True).sheetnames
        for sheet in sheets:
            table1 = pd.read_excel(excel_file, sheet_name=sheet, header=0)
            for r in range(0, table1.index.stop):
                srow = student()
                srow.student_group = sheet
                srow.student_name = table1.iat[r, 1]
                srow.student_ID = table1.iat[r, 2]
                srow.student_rn = table1.iat[r, 3]
                srow.student_phone = table1.iat[r, 4]
                srow.detail_address = table1.iat[r, 5]
                srow.board_ID = post_id
                slist.append(srow)

    context = {
        'students': slist,
        'board_id': post_id,
        'img_cnt': 0,
    }

    return render(request, 'board_register.html', context)


# 게시판 글쓰기
def board_insert(request):
    title = request.POST['title']
    school_name = request.session['school_name']
    school_id = request.session['school_ID']
    post_id = request.POST['board_id2']

    # 학생정보 업데이트
    iname = request.POST.getlist('inputname')
    iid = request.POST.getlist('inputid')
    irn = request.POST.getlist('inputrn')
    iadd = request.POST.getlist('inputadd')
    ipn = request.POST.getlist('inputpn')
    igroup = request.POST.getlist('inputgroup')

    ilist = []
    if len(iname) > 0:
        img_saver = student.objects.exclude(Q(student_img__isnull=True) | Q(student_img__exact=''))
        if img_saver.exists():
            for i in img_saver:
                ilist.append(i.student_img)

        for i in range(len(iname)):
            # 학생정보 입력 error
            if len(iname[i]) == 0 or len(irn[i]) == 0 or len(ipn[i]) == 0 \
                    or len(iid[i]) == 0 or len(iadd[i]) == 0:
                messages.warning(request, '학생필수정보가 입력되지 않은 행이 존재합니다.')
                return redirect(request.META.get('HTTP_REFERER'))

        students = student.objects.filter(board_ID=post_id)
        if students.exists():
            students.delete()

        for i in range(len(iname)):
            if student.objects.filter(Q(school_ID=school_id) & Q(student_ID=iid[i]) & Q(board_ID=post_id)).exists():
                student_row = student.objects.filter(
                    Q(school_ID=school_id) & Q(student_ID=iid[i]) & Q(board_ID=post_id))
                student_row.update(
                    student_group=igroup[i],
                    student_name=iname[i],
                    student_rn=irn[i],
                    student_phone=ipn[i],
                    detail_address=iadd[i],
                )
            else:
                student_row = student.objects.create(
                    school_ID=school_id,
                    student_ID=iid[i],
                    student_group=igroup[i],
                    student_name=iname[i],
                    student_rn=irn[i],
                    student_phone=ipn[i],
                    board_ID=post_id,
                    detail_address=iadd[i],
                )

    # 학생이미지
    if request.FILES.getlist('file2') is not None:
        for uploaded_file in request.FILES.getlist('file2'):
            name_org = uploaded_file.name
            fs = FileSystemStorage(location='static/board/image')
            name = fs.save(name_org, uploaded_file.file)
            filename = name_org.split('.')[0].split('_')
            print(filename)
            students = student.objects.filter(
                Q(student_name=filename[1]) & Q(student_ID=filename[0]) & Q(board_ID=post_id)
            )
            if students.exists():
                student_row = student.objects.get(
                    Q(student_name=filename[1]) & Q(student_ID=filename[0]) & Q(board_ID=post_id)
                )
                student_row.student_img = name_org
                student_row.save()
    if len(ilist) > 0:
        for i in ilist:
            filename = i.split('.')[0].split('_')
            print(filename)
            students = student.objects.filter(
                Q(student_name=filename[1]) & Q(student_ID=filename[0]) & Q(board_ID=post_id)
            )
            if students.exists():
                student_row = student.objects.get(
                    Q(student_name=filename[1]) & Q(student_ID=filename[0]) & Q(board_ID=post_id)
                )
                student_row.student_img = i
                student_row.save()

    if title != "":
        if Board.objects.filter(id=post_id).exists():
            row = Board.objects.get(id=post_id)
            row.title = title
            row.save()
        else:
            rows = Board.objects.create(title=title, school_name=school_name)
        return redirect('/board')
    else:
        return redirect('/board_write')


# 게시판 상세조회
def board_view(request):
    post_id = request.GET.get('board_id')
    school_id = request.session['school_ID']
    school_name = request.session['school_name']
    boards = get_object_or_404(Board, id=post_id)
    school = get_object_or_404(school_info, school_ID=school_id)

    # 조회 권한
    if school_id != 'admin' and boards.school_name != school_name:
        messages.warning(request, '해당 학교만 조회가 가능합니다')
        return redirect('/board')

    students = student.objects.filter(board_ID=post_id)
    final_yn = boards.final_yn

    if school_id == boards.school_name:
        board_auth = True
    else:
        board_auth = False

    context = {
        'boards': boards,
        'board_auth': board_auth,
        'board_id': post_id,
        'students': students,
        'school': school,
        'final_yn': final_yn,
    }

    response = render(request, 'board_template.html', context)

    boards.hit_count += 1
    boards.save()

    return response


# 게시글 수정
def board_edit(request):
    post_id = request.GET['board_id']
    boards = get_object_or_404(Board, id=post_id)
    students = student.objects.filter(board_ID=post_id)
    img_cnt = 0
    if students.exists():
        img_st = students.exclude(Q(student_img__isnull=True) | Q(student_img__exact=''))
        img_cnt = img_st.count()

    if request.FILES.get('excelfile') is not None:
        students = []
        excel_file = request.FILES.get('excelfile')
        fs = FileSystemStorage(location='static/board/xlsx')
        name = fs.save(excel_file.name, excel_file)

        sheets = load_workbook(excel_file, read_only=True).sheetnames
        for sheet in sheets:
            table1 = pd.read_excel(excel_file, sheet_name=sheet, header=0)
            for r in range(0, table1.index.stop):
                srow = student()
                srow.student_group = sheet
                srow.student_name = table1.iat[r, 1]
                srow.student_ID = table1.iat[r, 2]
                srow.student_rn = table1.iat[r, 3]
                srow.student_phone = table1.iat[r, 4]
                srow.detail_address = table1.iat[r, 5]
                srow.board_ID = post_id
                students.append(srow)
    context = {
        'boards': boards,
        'board_id': post_id,
        'students': students,
        'img_cnt': img_cnt,
    }

    return render(request, 'board_register.html', context)


# 게시글 첨부파일 다운로드 한글명 인코딩
def board_download_view(request):
    post_id = request.GET['board_id']
    boards = get_object_or_404(Board, id=post_id)
    url = boards.upload_files.url[1:]
    print(type(url))
    print(url)
    file_url = urllib.parse.unquote(url)

    if os.path.exists(file_url):
        with open(file_url, 'rb') as fh:
            # quote_file_url = urllib.parse.quote(file_url.encode('utf-8'))
            quote_file_url = urllib.parse.quote(boards.filename.encode('utf-8'))
            response = HttpResponse(fh.read(), content_type=mimetypes.guess_type(file_url)[0])
            # response['Content-Disposition'] = 'attachment;filename*=UTF-8\'\'%s' % quote_file_url[29:]
            response['Content-Disposition'] = 'attachment;filename*=UTF-8\'\'%s' % quote_file_url
            return response
        raise Http404


# 게시글 최종완료
def board_final(request):
    post_id = request.GET['board_id']
    boards = Board.objects.get(id=post_id)
    boards.final_yn = 1
    boards.save()

    return redirect('/board')